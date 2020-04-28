from django.http import HttpResponse
from django.shortcuts import render
import openpyxl
#import importer as imp
from .import models
NO = 0
TITLE = 1
INVENTORS = 2
APPLICANTS = 3
PUBLICATION_NUMBER = 4
COUNTRY = 5
EARLIEST_PRIORITY = 6
ipc = 7
cpc = 8
PUBLICATION_DATE = 9
PUBLICATION_YEAR = 10
EARLIEST_PUBLICATION = 11
FAMILY_NUMBER = 12

def index(request):
    if "GET" == request.method:
        return render(request, 'index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size


        print('teekshan')
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb.active


        # getting a particular sheet by name out of many sheets
        #worksheet = wb["Sheet1"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        rownum = 0
        for row in worksheet.iter_rows():
            if rownum == 0:
                rownum += 1
            else:
                the_row = models.report( No =row[NO], Title=row[TITLE], Inventors=row[INVENTORS], Applicants=row[APPLICANTS],
                                         Publication_number=row[PUBLICATION_NUMBER], Country=row[COUNTRY], Earliest_priority=row[EARLIEST_PRIORITY],
                                         IPC=row[ipc],CPC=row[cpc], Publication_date=row[PUBLICATION_DATE], Publication_Year=row[PUBLICATION_YEAR],
                                         Earliest_publication=row[EARLIEST_PUBLICATION],Family_number=row[FAMILY_NUMBER]
                                         )

                the_row.save()

        return render(request, 'index.html', {"excel_data":excel_data})

#def importer(request):
  #  dfile = openpyxl.load_workbook(excel_file)
    # reader=csv.reader(dfile)
  #  print(teekshan)

  #  rownum = 0
   # for row in reader:
    #    if rownum == 0:
    #        rownum += 1
    #    else:
    #        the_row = models.report(title=row[0], Inventors=row[1], Applicants=row[2], Publication_number=row[3],
     #                            Country=row[4],
     #                            Earliest_priority=row[5], IPC=row[6], CPC=row[7], Publication_date=row[8],
      #                           Publication_year=row[9],
     #                            Earliest_publication=row[10], Family_number=row[11])

      #      the_row.save()
    #imp.get_data = ("excel_file")
   # return render(request,'index.html')
